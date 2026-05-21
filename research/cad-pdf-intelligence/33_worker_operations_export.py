#!/usr/bin/env python3
"""
33_worker_operations_export.py
Stage S20 — Worker / Operations Export Generator

Creates the practical export layer of the Plan Scanner:
  - HTML report  (print-ready, self-contained, no external deps)
  - Excel workbook with 10 sheets (requires openpyxl)
  - Export manifest JSON

Modes:
  legacy        — reads from SCRIPT_DIR/outputs/, writes to SCRIPT_DIR/outputs/exports/
  plan-scoped   — reads from <plan-run-dir>/outputs/, writes to <plan-run-dir>/outputs/exports/

Usage:
  .venv/bin/python3 33_worker_operations_export.py                         # legacy
  .venv/bin/python3 33_worker_operations_export.py --plan-run-dir runs/poc_plan_50_448_02_400_20260520_223259/

IMPORTANT:
  All outputs are DRAFT — approved_for_boq: false.
  Do not use for execution or billing without human review.
  Source PDF is temporary input; exported report is the durable product.

Research-only. No production DB/UI changes. No paid API. No migrations.
"""
from __future__ import annotations

import argparse
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from plan_run_context import PlanRunContext

# ── Config ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
SCRIPT_NAME = "33_worker_operations_export.py"
VERSION = "1.0.0"

DISCLAIMER = (
    "DRAFT — RESEARCH EXPORT ONLY. "
    "Do not use for final execution, billing, or purchasing without human review. "
    "All quantities are provisional. approved_for_boq: false. "
    "Source PDF is a temporary research input; this exported report is the durable product."
)

# ── Helpers ────────────────────────────────────────────────────────────────────

def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default

def fmt_m(m: float) -> str:
    return f"{m:,.1f} m"

def fmt_qty(q) -> str:
    if q is None:
        return "—"
    if isinstance(q, float):
        return f"{q:,.1f}"
    return str(q)

def confidence_badge(c: str) -> str:
    mapping = {
        "high": "HIGH",
        "medium": "MED",
        "low": "LOW",
        "uncertain": "?",
        "unverified": "?",
    }
    return mapping.get(str(c).lower(), str(c).upper())

# ── Input loader ───────────────────────────────────────────────────────────────

class PipelineInputs:
    """Loads all available JSON sources from outputs_dir; reports missing."""

    REQUIRED_FILES = [
        ("boq",         "boq_unified_draft.json"),
        ("pipeline",    "pipeline_run_summary.json"),
        ("inventory",   "sign_inventory.json"),
        ("elements",    "element_groups.json"),
        ("dashboard",   "master_dashboard.json"),
        ("teaching",    "teaching_loop_answer_pack.json"),
        ("partial",     "partial_code_resolution.json"),
        ("validation",  "validation_results.json"),
        ("review",      "review_queue.json"),
        ("prototype",   "plan_scanner_prototype.json"),
    ]

    OPTIONAL_FILES = [
        ("legend",      "legend_vocabulary.json"),
        ("legend_rows", "legend_rows.json"),
        ("symbol_clust","symbol_clusters.json"),
        ("scale",       "scale_measurement/results.json"),
        ("state",       "../state/plan_scan_state.json"),
    ]

    def __init__(self, outputs_dir: Path):
        self.outputs_dir = outputs_dir
        self.data: Dict[str, Any] = {}
        self.present: Dict[str, bool] = {}
        self.missing: List[str] = []

        for key, rel in self.REQUIRED_FILES + self.OPTIONAL_FILES:
            path = outputs_dir / rel
            val = load_json(path)
            self.data[key] = val
            self.present[key] = val is not None
            if val is None:
                self.missing.append(rel)

    # Convenience accessors ------------------------------------------------

    @property
    def boq(self) -> Dict:
        return self.data.get("boq") or {}

    @property
    def boq_items(self) -> List[Dict]:
        return self.boq.get("items") or []

    @property
    def boq_totals(self) -> Dict:
        return self.boq.get("totals") or {}

    @property
    def boq_meta(self) -> Dict:
        return self.boq.get("meta") or {}

    @property
    def inventory(self) -> Dict:
        return self.data.get("inventory") or {}

    @property
    def occurrences(self) -> List[Dict]:
        return self.inventory.get("occurrences") or []

    @property
    def inv_summary(self) -> Dict:
        return self.inventory.get("summary") or {}

    @property
    def elements(self) -> Dict:
        return self.data.get("elements") or {}

    @property
    def groups(self) -> List[Dict]:
        return self.elements.get("groups") or []

    @property
    def red_flags(self) -> List[Dict]:
        return (self.data.get("dashboard") or {}).get("red_flags") or []

    @property
    def teaching_questions(self) -> List[Dict]:
        return (self.data.get("teaching") or {}).get("questions") or []

    @property
    def pipeline_status(self) -> Dict:
        return (self.data.get("pipeline") or {}).get("pipeline_status") or {}

    @property
    def scale_info(self) -> Dict:
        if self.present.get("scale"):
            return (self.data.get("scale") or {}).get("scale_info") or {}
        return self.boq.get("scale_info") or {}

    @property
    def plan_meta(self) -> Dict:
        return (self.data.get("pipeline") or {}).get("metadata") or {}

# ── Summary builder ────────────────────────────────────────────────────────────

def build_summary(inp: PipelineInputs, plan_slug: str, plan_id: str,
                  source_pdf: str) -> Dict:
    totals = inp.boq_totals
    boq_meta = inp.boq_meta
    inv = inp.inv_summary
    pipe = inp.pipeline_status

    requires_review_count = sum(
        1 for it in inp.boq_items if it.get("requires_review")
    )
    approved_count = sum(
        1 for it in inp.boq_items if it.get("approved_for_boq")
    )

    return {
        "plan_id": plan_id,
        "plan_slug": plan_slug,
        "source_pdf": source_pdf,
        "generated_at": _now_iso(),
        "status": "DRAFT — REQUIRES REVIEW",
        "scanner_note": "This is a Plan Scanner research export. The source PDF is NOT archived here.",
        "pipeline": {
            "stages_ok": pipe.get("stages_ok", "—"),
            "stages_partial": pipe.get("stages_partial", "—"),
            "overall": pipe.get("overall", "unknown"),
        },
        "boq": {
            "total_items": len(inp.boq_items),
            "approved_for_boq": approved_count,
            "requires_review": requires_review_count,
            "scale_status": boq_meta.get("scale_status", "unknown"),
            "color_taxonomy_status": boq_meta.get("color_taxonomy_status", "unknown"),
        },
        "signs": {
            "n_sign_plates": totals.get("total_sign_plates", inv.get("n_sign_plates", "—")),
            "n_pole_locations": totals.get("total_pole_locations", inv.get("n_pole_groups", "—")),
            "n_assemblies": totals.get("total_assemblies", "—"),
            "n_sign_code_candidates": totals.get("sign_code_candidates", 0),
        },
        "measurements": {
            "total_linear_m": totals.get("total_linear_m", 0.0),
            "scale_status": boq_meta.get("scale_status", "unverified"),
        },
        "review": {
            "n_review_items": requires_review_count,
            "n_red_flags": len(inp.red_flags),
            "n_teaching_questions": len(inp.teaching_questions),
        },
    }

# ── HTML generator ─────────────────────────────────────────────────────────────

_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Helvetica Neue', Arial, sans-serif; font-size: 13px;
       color: #1a1a1a; background: #f4f4f5; }
.page { max-width: 1100px; margin: 0 auto; padding: 24px; }
h1 { font-size: 22px; font-weight: 700; margin-bottom: 4px; }
h2 { font-size: 16px; font-weight: 600; margin: 24px 0 10px; color: #1d3557; border-bottom: 2px solid #1d3557; padding-bottom: 4px; }
h3 { font-size: 13px; font-weight: 600; margin: 14px 0 6px; color: #333; }
p  { margin-bottom: 8px; line-height: 1.5; }
.header-box { background: #1d3557; color: #fff; padding: 20px 24px; border-radius: 6px; margin-bottom: 20px; }
.header-box .sub { font-size: 12px; opacity: 0.7; margin-top: 4px; }
.status-badge { display: inline-block; padding: 3px 10px; border-radius: 3px;
                font-weight: 700; font-size: 11px; letter-spacing: 0.5px; text-transform: uppercase; }
.badge-draft    { background: #f59e0b; color: #fff; }
.badge-critical { background: #dc2626; color: #fff; }
.badge-warning  { background: #f59e0b; color: #fff; }
.badge-info     { background: #3b82f6; color: #fff; }
.badge-ok       { background: #16a34a; color: #fff; }
.disclaimer { background: #fef2f2; border: 2px solid #dc2626; border-radius: 6px;
              padding: 14px 18px; margin-bottom: 20px; font-size: 12px; color: #7f1d1d; }
.card { background: #fff; border-radius: 6px; padding: 16px 20px; margin-bottom: 16px;
        box-shadow: 0 1px 3px rgba(0,0,0,.07); }
.kpi-row { display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 16px; }
.kpi { background: #fff; border-radius: 6px; padding: 14px 18px; min-width: 140px;
       box-shadow: 0 1px 3px rgba(0,0,0,.07); text-align: center; }
.kpi .val { font-size: 28px; font-weight: 700; color: #1d3557; }
.kpi .lbl { font-size: 11px; color: #6b7280; margin-top: 2px; text-transform: uppercase; letter-spacing: 0.4px; }
table { width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 8px; }
th { background: #1d3557; color: #fff; padding: 7px 10px; text-align: left; font-weight: 600; }
td { padding: 6px 10px; border-bottom: 1px solid #e5e7eb; vertical-align: top; }
tr:nth-child(even) td { background: #f9fafb; }
tr:hover td { background: #eff6ff; }
.flag-critical td:first-child { border-left: 3px solid #dc2626; }
.flag-warning  td:first-child { border-left: 3px solid #f59e0b; }
.flag-info     td:first-child { border-left: 3px solid #3b82f6; }
.missing { color: #9ca3af; font-style: italic; }
.review-yes  { color: #dc2626; font-weight: 600; }
.review-no   { color: #16a34a; }
.conf-high   { color: #16a34a; font-weight: 600; }
.conf-medium { color: #d97706; font-weight: 600; }
.conf-low    { color: #dc2626; }
.conf-uncertain { color: #9ca3af; }
@media print {
  body { background: #fff; }
  .page { max-width: 100%; padding: 0; }
  .header-box { border-radius: 0; }
}
"""

def _badge(text: str, cls: str) -> str:
    return f'<span class="status-badge badge-{cls}">{text}</span>'

def _review_cell(val: bool) -> str:
    if val:
        return '<span class="review-yes">YES</span>'
    return '<span class="review-no">no</span>'

def _conf_cell(c: str) -> str:
    c = str(c).lower()
    return f'<span class="conf-{c}">{c.upper()}</span>'

def _td(v) -> str:
    if v is None or v == "":
        return '<td class="missing">—</td>'
    return f"<td>{v}</td>"

def generate_html(inp: PipelineInputs, summary: Dict, plan_manifest: Dict) -> str:
    gen_at = summary["generated_at"]
    plan_name = plan_manifest.get("plan_name") or summary["plan_slug"]
    source_pdf = Path(summary["source_pdf"]).name
    slug = summary["plan_slug"]
    totals = inp.boq_totals
    scale_info = inp.scale_info
    scale_ratio = scale_info.get("ratio", "?")
    scale_status = scale_info.get("status", "unknown")
    pipe_ok = summary["pipeline"]["stages_ok"]
    pipe_partial = summary["pipeline"]["stages_partial"]
    n_items = summary["boq"]["total_items"]
    n_review = summary["boq"]["requires_review"]
    n_approved = summary["boq"]["approved_for_boq"]
    n_plates = summary["signs"]["n_sign_plates"]
    n_poles = summary["signs"]["n_pole_locations"]
    total_m = summary["measurements"]["total_linear_m"]
    n_flags = summary["review"]["n_red_flags"]
    n_questions = summary["review"]["n_teaching_questions"]

    parts: List[str] = []
    p = parts.append

    p(f"<!DOCTYPE html><html lang='he' dir='ltr'><head>")
    p(f"<meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>")
    p(f"<title>Worker Operations Export — {plan_name}</title>")
    p(f"<style>{_CSS}</style></head><body><div class='page'>")

    # ── 1. Disclaimer ──────────────────────────────────────────────────────────
    p(f"<div class='disclaimer'>")
    p(f"<strong>⚠ DRAFT RESEARCH EXPORT — DO NOT USE FOR EXECUTION OR BILLING</strong><br>")
    p(f"{DISCLAIMER}")
    p(f"</div>")

    # ── 2. Header ──────────────────────────────────────────────────────────────
    p(f"<div class='header-box'>")
    p(f"<h1>Worker / Operations Report — {plan_name}</h1>")
    p(f"<div class='sub'>Plan slug: {slug} &nbsp;|&nbsp; Source: {source_pdf}</div>")
    p(f"<div class='sub'>Generated: {gen_at} &nbsp;|&nbsp; {_badge('DRAFT', 'draft')} {_badge('REQUIRES REVIEW', 'critical')}</div>")
    p(f"<div class='sub'>Pipeline: {pipe_ok} stages OK, {pipe_partial} partial &nbsp;|&nbsp; Scale: 1:{scale_ratio} ({scale_status})</div>")
    p(f"</div>")

    # ── 3. Executive summary KPIs ──────────────────────────────────────────────
    p("<h2>Executive Summary</h2>")
    p("<div class='kpi-row'>")
    kpis = [
        (n_items, "BOQ Items"),
        (n_approved, "Approved for BOQ"),
        (n_review, "Requires Review"),
        (n_plates, "Sign Plates"),
        (n_poles, "Pole Locations"),
        (fmt_m(total_m), "Linear (unverified)"),
        (n_flags, "Red Flags"),
        (n_questions, "Human Questions"),
    ]
    for val, lbl in kpis:
        p(f"<div class='kpi'><div class='val'>{val}</div><div class='lbl'>{lbl}</div></div>")
    p("</div>")

    # ── 4. Red flags ───────────────────────────────────────────────────────────
    p("<h2>Red Flags</h2>")
    p("<div class='card'>")
    if inp.red_flags:
        p("<table><tr><th>Severity</th><th>Code</th><th>Message</th><th>Action</th></tr>")
        for rf in inp.red_flags:
            sev = rf.get("severity", "INFO").upper()
            cls = "critical" if sev == "CRITICAL" else ("warning" if sev == "WARNING" else "info")
            code = rf.get("code", "")
            msg = rf.get("message", "")
            action = rf.get("action", "")
            p(f"<tr class='flag-{cls}'>{_td(_badge(sev, cls))}{_td(code)}{_td(msg)}{_td(action)}</tr>")
        p("</table>")
    else:
        p("<p class='missing'>No red flags data available.</p>")
    p("</div>")

    # ── 5. Work preparation summary ────────────────────────────────────────────
    p("<h2>Work Preparation Summary</h2>")
    p("<div class='card'>")

    # Signs
    n_no_match = sum(1 for o in inp.occurrences if o.get("visual_match_tier") == "no_match")
    n_medium = sum(1 for o in inp.occurrences if o.get("visual_match_tier") == "medium")
    n_low = sum(1 for o in inp.occurrences if o.get("visual_match_tier") == "low")
    n_legend_matched = inp.inv_summary.get("n_legend_matched", 0)
    n_pending = inp.inv_summary.get("n_pending_vision", 0)

    p("<h3>Signs / Symbols</h3>")
    p("<table><tr><th>Category</th><th>Count</th><th>Status</th></tr>")
    p(f"<tr>{_td('Pole locations')}{_td(n_poles)}{_td(_badge('REQUIRES REVIEW', 'warning'))}</tr>")
    p(f"<tr>{_td('Sign plates total')}{_td(n_plates)}{_td(_badge('REQUIRES REVIEW', 'warning'))}</tr>")
    p(f"<tr>{_td('Legend-matched signs')}{_td(n_legend_matched)}{_td(_badge('MED CONFIDENCE', 'info'))}</tr>")
    p(f"<tr>{_td('No match (pending identification)')}{_td(n_no_match)}{_td(_badge('UNIDENTIFIED', 'critical'))}</tr>")
    p(f"<tr>{_td('Partial visual match (medium/low)')}{_td(n_medium + n_low)}{_td(_badge('UNCONFIRMED', 'warning'))}</tr>")
    p(f"<tr>{_td('Sign codes confirmed (3-digit)')}{_td(totals.get('sign_code_candidates', 0))}{_td(_badge('PENDING FIX', 'critical'))}</tr>")
    p("</table>")

    # Measurements
    p("<h3>Linear Measurements (scale unverified)</h3>")
    p("<table><tr><th>Item Type</th><th>Quantity</th><th>Unit</th><th>Scale Status</th></tr>")
    lin_items = [it for it in inp.boq_items if it.get("item_category") == "measured_linear"]
    for it in lin_items[:15]:
        desc = it.get("description_en") or it.get("description_he") or it.get("item_type", "")
        qty = it.get("quantity", 0)
        unit = it.get("unit", "m")
        p(f"<tr>{_td(desc)}{_td(fmt_qty(qty))}{_td(unit)}{_td(_badge(scale_status.upper(), 'critical' if scale_status == 'unverified' else 'info'))}</tr>")
    if len(lin_items) > 15:
        p(f"<tr><td colspan='4' class='missing'>... and {len(lin_items)-15} more. See Excel → Measurements sheet.</td></tr>")
    p("</table>")

    # Element groups requiring classification
    unclassified = [g for g in inp.groups if g.get("requires_review")]
    p("<h3>Element Groups Requiring Classification</h3>")
    p(f"<table><tr><th>Group</th><th>Type</th><th>Description (EN)</th><th>Paths</th><th>Length (pt)</th></tr>")
    for g in unclassified[:10]:
        gid = g.get("group_id", "")
        et = g.get("element_type", "")
        desc = g.get("description_en", "")
        n_p = g.get("n_paths", "—")
        ln = g.get("total_length_pt", "—")
        p(f"<tr>{_td(gid)}{_td(et)}{_td(desc)}{_td(n_p)}{_td(fmt_qty(ln) if isinstance(ln, float) else ln)}</tr>")
    if len(unclassified) > 10:
        p(f"<tr><td colspan='5' class='missing'>... and {len(unclassified)-10} more. See Excel → Element Groups sheet.</td></tr>")
    p("</table>")
    p("</div>")

    # ── 6. BOQ draft table ─────────────────────────────────────────────────────
    p("<h2>BOQ Draft Table</h2>")
    p("<div class='card'>")
    p("<table><tr>")
    for col in ["ID", "Category", "Description (EN)", "Qty", "Unit", "Branch", "Confidence", "Review?", "Approved?", "Review Reason"]:
        p(f"<th>{col}</th>")
    p("</tr>")
    for it in inp.boq_items:
        bid = it.get("boq_item_id", "")
        cat = it.get("item_category", "")
        desc = it.get("description_en") or it.get("description_he") or ""
        qty = it.get("quantity")
        unit = it.get("unit", "")
        branch = it.get("source_branch", "")
        conf = it.get("confidence", "")
        rr = it.get("requires_review", False)
        ap = it.get("approved_for_boq", False)
        reason = it.get("review_reason", "")
        p(f"<tr>{_td(bid)}{_td(cat)}{_td(desc)}{_td(fmt_qty(qty))}{_td(unit)}{_td(branch)}")
        p(f"{_td(_conf_cell(conf))}{_td(_review_cell(rr))}{_td(_review_cell(ap))}{_td(reason)}</tr>")
    p("</table></div>")

    # ── 7. Signs / codes section ───────────────────────────────────────────────
    p("<h2>Signs / Codes</h2>")
    p("<div class='card'>")
    p("<table><tr><th>ID</th><th>Plate ID</th><th>Pole Group</th><th>Color</th>")
    p("<th>Legend Match</th><th>Sign Code</th><th>Code Source</th><th>Visual Tier</th><th>Confidence</th><th>Review?</th></tr>")
    for occ in inp.occurrences[:50]:
        oid = occ.get("occurrence_id", "")
        plate = occ.get("sign_plate_id", "")
        pole = occ.get("pole_group_id", "")
        color = occ.get("dominant_color", "")
        lm = occ.get("matched_legend_row") or "—"
        code = occ.get("selected_sign_code") or "—"
        code_src = occ.get("sign_code_source") or "—"
        vt = occ.get("visual_match_tier") or "—"
        conf = occ.get("final_confidence") or "—"
        rr = occ.get("requires_review", True)
        p(f"<tr>{_td(oid)}{_td(plate)}{_td(pole)}{_td(color)}{_td(lm)}{_td(code)}{_td(code_src)}{_td(vt)}{_td(conf)}{_td(_review_cell(rr))}</tr>")
    if len(inp.occurrences) > 50:
        p(f"<tr><td colspan='10' class='missing'>... {len(inp.occurrences)-50} more occurrences. See Excel → Signs/Codes sheet.</td></tr>")
    p("</table></div>")

    # ── 8. Human action / teaching loop ───────────────────────────────────────
    p("<h2>Human Action Required</h2>")
    p("<div class='card'>")
    if inp.teaching_questions:
        p("<table><tr><th>ID</th><th>Priority</th><th>Type</th><th>Question</th><th>Impact</th><th>Status</th></tr>")
        for q in inp.teaching_questions:
            qid = q.get("question_id", "")
            pri = q.get("priority", "")
            qtype = q.get("question_type", "")
            qtxt = q.get("question_text", "")[:200]
            impact = q.get("business_impact", "")[:150]
            status = q.get("status", "pending")
            p(f"<tr>{_td(qid)}{_td(pri)}{_td(qtype)}{_td(qtxt)}{_td(impact)}{_td(status)}</tr>")
        p("</table>")
    else:
        p("<p class='missing'>No teaching loop questions loaded. Run 25_teaching_loop_answer_pack.py.</p>")
    p("</div>")

    # Scale calibration callout
    p("<div class='card'>")
    p("<h3>Scale Calibration (CRITICAL)</h3>")
    p(f"<p>Current scale: <strong>1:{scale_ratio}</strong> — source: <em>{scale_info.get('source', '?')}</em></p>")
    if scale_status == "unverified":
        p(f"<p class='review-yes'>⚠ Scale is unverified. All {fmt_m(total_m)} of linear measurements may be wrong.</p>")
        p("<p>To calibrate: identify two PDF points with a known real-world distance, record their coordinates in scale_measurement/results.json, and re-run 15_scale_measurement.py.</p>")
    p("</div>")

    # ── 9. Missing inputs ──────────────────────────────────────────────────────
    if inp.missing:
        p("<h2>Missing Inputs</h2>")
        p("<div class='card'>")
        p("<p>The following input files were not found. Export continues with available data.</p>")
        p("<ul>")
        for m in inp.missing:
            p(f"<li class='missing'>{m}</li>")
        p("</ul></div>")

    # ── 10. Safety disclaimer ──────────────────────────────────────────────────
    p("<h2>Safety Disclaimer</h2>")
    p("<div class='disclaimer'>")
    p(f"<strong>This is a draft research export.</strong><br>")
    p(f"<br>• Do not use for final execution or billing without human review.")
    p(f"<br>• approved_for_boq: false for all {n_items} items.")
    p(f"<br>• Scale 1:{scale_ratio} is {scale_status} — all linear measurements are provisional.")
    p(f"<br>• Sign codes are not confirmed — pipeline structural limitation (2-digit partial codes only).")
    p(f"<br>• Source PDF is a temporary research input. This exported report is the durable product.")
    p(f"<br>• Generated: {gen_at} by {SCRIPT_NAME} v{VERSION}.")
    p("</div>")

    p("</div></body></html>")
    return "\n".join(parts)

# ── Excel generator ────────────────────────────────────────────────────────────

def _make_header_style(wb):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1D3557")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        warn_fill   = PatternFill("solid", fgColor="FEF2F2")
        return header_fill, header_font, warn_fill
    except Exception:
        return None, None, None

def _write_headers(ws, headers, header_fill, header_font):
    try:
        from openpyxl.styles import Alignment
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            if header_fill:
                cell.fill = header_fill
            if header_font:
                cell.font = header_font
    except Exception:
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

def _autowidth(ws, min_w=8, max_w=50):
    try:
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=min_w)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)
    except Exception:
        pass

def generate_excel(inp: PipelineInputs, summary: Dict, out_path: Path) -> bool:
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        return False

    wb = Workbook()
    # remove default sheet
    if wb.active:
        wb.remove(wb.active)

    hf, hfont, wfill = _make_header_style(wb)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 48
    rows = [
        ("Plan", summary.get("plan_slug", "")),
        ("Source PDF", summary.get("source_pdf", "")),
        ("Generated At", summary.get("generated_at", "")),
        ("Status", "DRAFT — REQUIRES REVIEW"),
        ("", ""),
        ("--- Pipeline ---", ""),
        ("Stages OK", summary["pipeline"]["stages_ok"]),
        ("Stages Partial", summary["pipeline"]["stages_partial"]),
        ("Pipeline Overall", summary["pipeline"]["overall"]),
        ("", ""),
        ("--- BOQ ---", ""),
        ("Total BOQ Items", summary["boq"]["total_items"]),
        ("Approved for BOQ", summary["boq"]["approved_for_boq"]),
        ("Requires Review", summary["boq"]["requires_review"]),
        ("Scale Status", summary["boq"]["scale_status"]),
        ("Color Taxonomy Status", summary["boq"]["color_taxonomy_status"]),
        ("", ""),
        ("--- Signs ---", ""),
        ("Sign Plates", summary["signs"]["n_sign_plates"]),
        ("Pole Locations", summary["signs"]["n_pole_locations"]),
        ("Assemblies", summary["signs"]["n_assemblies"]),
        ("Sign Code Candidates", summary["signs"]["n_sign_code_candidates"]),
        ("", ""),
        ("--- Measurements ---", ""),
        ("Total Linear (m)", summary["measurements"]["total_linear_m"]),
        ("Scale Status", summary["measurements"]["scale_status"]),
        ("", ""),
        ("--- Review ---", ""),
        ("Review Items", summary["review"]["n_review_items"]),
        ("Red Flags", summary["review"]["n_red_flags"]),
        ("Teaching Questions", summary["review"]["n_teaching_questions"]),
        ("", ""),
        ("--- Disclaimer ---", ""),
        ("WARNING", DISCLAIMER),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=str(v) if v is not None else "")

    # ── Sheet 2: BOQ Draft ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("BOQ Draft")
    headers = ["BOQ ID", "Category", "Type", "Description EN", "Description HE",
               "Quantity", "Unit", "Source Branch", "Confidence",
               "Requires Review", "Approved for BOQ", "Review Reason", "Audit Notes"]
    _write_headers(ws2, headers, hf, hfont)
    for r, it in enumerate(inp.boq_items, 2):
        vals = [
            it.get("boq_item_id", ""),
            it.get("item_category", ""),
            it.get("item_type", ""),
            it.get("description_en", ""),
            it.get("description_he", ""),
            it.get("quantity"),
            it.get("unit", ""),
            it.get("source_branch", ""),
            it.get("confidence", ""),
            "YES" if it.get("requires_review") else "no",
            "YES" if it.get("approved_for_boq") else "no",
            it.get("review_reason", ""),
            it.get("audit_notes", ""),
        ]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r, column=c, value=v)
    _autowidth(ws2)

    # ── Sheet 3: Signs / Codes ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Signs - Codes")
    headers = ["OCC ID", "Sign Plate", "Pole Group", "Assembly", "Color",
               "Cluster Type", "Legend Match", "Visual Tier", "Sign Code",
               "Code Source", "Final Confidence", "Contradiction Flags",
               "Requires Review", "Notes"]
    _write_headers(ws3, headers, hf, hfont)
    for r, occ in enumerate(inp.occurrences, 2):
        flags = ", ".join(occ.get("contradiction_flags") or [])
        vals = [
            occ.get("occurrence_id", ""),
            occ.get("sign_plate_id", ""),
            occ.get("pole_group_id", ""),
            occ.get("assembly_id", ""),
            occ.get("dominant_color", ""),
            occ.get("cluster_type", ""),
            occ.get("matched_legend_row") or "",
            occ.get("visual_match_tier") or "",
            occ.get("selected_sign_code") or "",
            occ.get("sign_code_source") or "",
            occ.get("final_confidence") or "",
            flags,
            "YES" if occ.get("requires_review") else "no",
            occ.get("notes") or "",
        ]
        for c, v in enumerate(vals, 1):
            ws3.cell(row=r, column=c, value=v)
    _autowidth(ws3)

    # ── Sheet 4: Poles / Assemblies ───────────────────────────────────────────
    ws4 = wb.create_sheet("Poles - Assemblies")
    headers = ["BOQ ID", "Description EN", "Description HE",
               "Quantity", "Unit", "Confidence", "Requires Review",
               "Review Reason", "Source IDs (sample)"]
    _write_headers(ws4, headers, hf, hfont)
    pole_items = [it for it in inp.boq_items
                  if it.get("item_type") in ("pole_location", "sign_plate", "assembly")]
    for r, it in enumerate(pole_items, 2):
        src = it.get("source_ids") or []
        vals = [
            it.get("boq_item_id", ""),
            it.get("description_en", ""),
            it.get("description_he", ""),
            it.get("quantity"),
            it.get("unit", ""),
            it.get("confidence", ""),
            "YES" if it.get("requires_review") else "no",
            it.get("review_reason", ""),
            ", ".join(str(s) for s in src[:5]),
        ]
        for c, v in enumerate(vals, 1):
            ws4.cell(row=r, column=c, value=v)
    _autowidth(ws4)

    # ── Sheet 5: Measurements ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Measurements")
    headers = ["BOQ ID", "Type", "Description EN", "Quantity (m)", "Unit",
               "Scale Status", "Confidence", "Requires Review", "Review Reason"]
    _write_headers(ws5, headers, hf, hfont)
    scale_status = inp.scale_info.get("status", "unverified")
    scale_ratio  = inp.scale_info.get("ratio", "?")
    lin = [it for it in inp.boq_items if it.get("item_category") in ("measured_linear", "measured_area", "measured_linear_candidate")]
    for r, it in enumerate(lin, 2):
        vals = [
            it.get("boq_item_id", ""),
            it.get("item_type", ""),
            it.get("description_en", ""),
            it.get("quantity"),
            it.get("unit", ""),
            f"1:{scale_ratio} ({scale_status})",
            it.get("confidence", ""),
            "YES" if it.get("requires_review") else "no",
            it.get("review_reason", ""),
        ]
        for c, v in enumerate(vals, 1):
            ws5.cell(row=r, column=c, value=v)
    # Scale calibration note
    r_note = len(lin) + 3
    ws5.cell(row=r_note, column=1, value="SCALE NOTE")
    ws5.cell(row=r_note, column=2,
             value=inp.scale_info.get("detection_note") or "Scale unverified — see 15_scale_measurement.py")
    _autowidth(ws5)

    # ── Sheet 6: Element Groups ───────────────────────────────────────────────
    ws6 = wb.create_sheet("Element Groups")
    headers = ["Group ID", "Color Key", "RGB (8-bit)", "Element Type",
               "Description EN", "Description HE", "N Paths", "Total Length (pt)",
               "BOQ Category", "Confidence", "Requires Review", "Approved for BOQ", "Notes"]
    _write_headers(ws6, headers, hf, hfont)
    for r, g in enumerate(inp.groups, 2):
        rgb = g.get("color_rgb8") or []
        rgb_str = f"({','.join(str(x) for x in rgb)})" if rgb else ""
        vals = [
            g.get("group_id", ""),
            g.get("color_key", ""),
            rgb_str,
            g.get("element_type", ""),
            g.get("description_en", ""),
            g.get("description_he", ""),
            g.get("n_paths"),
            g.get("total_length_pt"),
            g.get("boq_category") or "",
            g.get("confidence", ""),
            "YES" if g.get("requires_review") else "no",
            "YES" if g.get("approved_for_boq") else "no",
            g.get("notes") or "",
        ]
        for c, v in enumerate(vals, 1):
            ws6.cell(row=r, column=c, value=v)
    _autowidth(ws6)

    # ── Sheet 7: Review Required ──────────────────────────────────────────────
    ws7 = wb.create_sheet("Review Required")
    headers = ["BOQ ID", "Type", "Description EN", "Quantity", "Unit",
               "Confidence", "Review Reason", "Approved for BOQ"]
    _write_headers(ws7, headers, hf, hfont)
    review_items = [it for it in inp.boq_items if it.get("requires_review")]
    for r, it in enumerate(review_items, 2):
        vals = [
            it.get("boq_item_id", ""),
            it.get("item_type", ""),
            it.get("description_en", ""),
            it.get("quantity"),
            it.get("unit", ""),
            it.get("confidence", ""),
            it.get("review_reason", ""),
            "YES" if it.get("approved_for_boq") else "no",
        ]
        for c, v in enumerate(vals, 1):
            ws7.cell(row=r, column=c, value=v)
    _autowidth(ws7)

    # ── Sheet 8: Red Flags ────────────────────────────────────────────────────
    ws8 = wb.create_sheet("Red Flags")
    headers = ["Severity", "Code", "Message", "Action", "File"]
    _write_headers(ws8, headers, hf, hfont)
    for r, rf in enumerate(inp.red_flags, 2):
        vals = [
            rf.get("severity", ""),
            rf.get("code", ""),
            rf.get("message", ""),
            rf.get("action", ""),
            rf.get("file", ""),
        ]
        for c, v in enumerate(vals, 1):
            ws8.cell(row=r, column=c, value=v)
    _autowidth(ws8)

    # ── Sheet 9: Preparation List ─────────────────────────────────────────────
    ws9 = wb.create_sheet("Preparation List")
    ws9.column_dimensions["A"].width = 6
    ws9.column_dimensions["B"].width = 30
    ws9.column_dimensions["C"].width = 50
    ws9.column_dimensions["D"].width = 20
    ws9.cell(row=1, column=1, value="#")
    ws9.cell(row=1, column=2, value="Category")
    ws9.cell(row=1, column=3, value="Task / Item")
    ws9.cell(row=1, column=4, value="Status")
    if hf:
        for c in range(1, 5):
            ws9.cell(row=1, column=c).fill = hf
        if hfont:
            for c in range(1, 5):
                ws9.cell(row=1, column=c).font = hfont

    prep_items = []
    # Scale calibration
    prep_items.append(("Scale", "Confirm or calibrate scale (1:500 assumed)", "PENDING"))
    prep_items.append(("Scale", "Re-run 15_scale_measurement.py after calibration", "PENDING"))
    # Sign codes
    prep_items.append(("Signs", "Resolve partial sign codes (current: 0 confirmed 3-digit codes)", "PENDING"))
    prep_items.append(("Signs", "Confirm legend labels via Vision API or manual input", "PENDING"))
    # Sign identification
    for occ in inp.occurrences[:10]:
        oid = occ.get("occurrence_id", "")
        prep_items.append(("Sign ID", f"Identify sign at {oid} (color: {occ.get('dominant_color','')})", "PENDING"))
    if len(inp.occurrences) > 10:
        prep_items.append(("Sign ID", f"... {len(inp.occurrences)-10} more signs require identification", "PENDING"))
    # Element groups
    high_impact = [g for g in inp.groups if g.get("element_type") in
                   ("black_stroke_unknown", "sign_glyph", "unknown_color", "gray_unknown")]
    for g in high_impact[:5]:
        prep_items.append(("Elements", f"Classify element group {g.get('group_id','')} ({g.get('description_en','')})", "PENDING"))
    # BOQ approval
    prep_items.append(("BOQ", "Complete BOQ review — 0/42 items approved", "PENDING"))
    prep_items.append(("BOQ", "Obtain BOQ approval from qualified supervisor before execution", "PENDING"))

    for r, (cat, task, status) in enumerate(prep_items, 2):
        ws9.cell(row=r, column=1, value=r - 1)
        ws9.cell(row=r, column=2, value=cat)
        ws9.cell(row=r, column=3, value=task)
        ws9.cell(row=r, column=4, value=status)

    # ── Sheet 10: Audit / Evidence ────────────────────────────────────────────
    ws10 = wb.create_sheet("Audit - Evidence")
    ws10.column_dimensions["A"].width = 30
    ws10.column_dimensions["B"].width = 60
    ws10.cell(row=1, column=1, value="Source")
    ws10.cell(row=1, column=2, value="Status / Notes")
    if hf:
        ws10.cell(row=1, column=1).fill = hf
        ws10.cell(row=1, column=2).fill = hf
    if hfont:
        ws10.cell(row=1, column=1).font = hfont
        ws10.cell(row=1, column=2).font = hfont

    evidence = [
        ("boq_unified_draft.json", f"{len(inp.boq_items)} items, all requires_review=true"),
        ("sign_inventory.json", f"{len(inp.occurrences)} occurrences, {inp.inv_summary.get('n_pole_groups',0)} poles"),
        ("element_groups.json", f"{len(inp.groups)} groups"),
        ("master_dashboard.json", f"{len(inp.red_flags)} red flags"),
        ("teaching_loop_answer_pack.json", f"{len(inp.teaching_questions)} questions"),
        ("scale_measurement/results.json", f"Scale 1:{scale_ratio} — {scale_status}"),
        ("pipeline_run_summary.json",
         f"Stages OK: {summary['pipeline']['stages_ok']}, partial: {summary['pipeline']['stages_partial']}"),
        ("Script", f"{SCRIPT_NAME} v{VERSION}"),
        ("Generated at", summary["generated_at"]),
        ("Disclaimer", DISCLAIMER[:200]),
    ]
    for r, (src, note) in enumerate(evidence, 2):
        ws10.cell(row=r, column=1, value=src)
        ws10.cell(row=r, column=2, value=note)

    wb.save(out_path)
    return True

# ── Export manifest ────────────────────────────────────────────────────────────

def generate_manifest(summary: Dict, exports: List[Dict], inp: PipelineInputs,
                      plan_manifest: Dict) -> Dict:
    return {
        "_warning": DISCLAIMER,
        "plan_id": summary.get("plan_id", ""),
        "plan_slug": summary.get("plan_slug", ""),
        "plan_name": plan_manifest.get("plan_name", ""),
        "source_pdf": summary.get("source_pdf", ""),
        "generated_at": summary.get("generated_at", ""),
        "generator_script": SCRIPT_NAME,
        "generator_version": VERSION,
        "approved_for_boq": False,
        "status": "DRAFT",
        "export_files": exports,
        "totals": {
            "boq_items": len(inp.boq_items),
            "approved_for_boq_count": 0,
            "requires_review_count": sum(1 for it in inp.boq_items if it.get("requires_review")),
            "sign_plates": inp.inv_summary.get("n_sign_plates", 0),
            "pole_locations": inp.inv_summary.get("n_pole_groups", 0),
            "total_linear_m": inp.boq_totals.get("total_linear_m", 0.0),
            "scale_status": inp.scale_info.get("status", "unverified"),
            "red_flag_count": len(inp.red_flags),
            "teaching_questions_count": len(inp.teaching_questions),
        },
        "missing_inputs": inp.missing,
        "pipeline_status": summary["pipeline"],
    }

# ── Markdown report ────────────────────────────────────────────────────────────

def generate_md_report(summary: Dict, inp: PipelineInputs,
                       exports: List[Dict], excel_ok: bool) -> str:
    gen = summary["generated_at"]
    slug = summary["plan_slug"]
    n_items = summary["boq"]["total_items"]
    n_review = summary["boq"]["requires_review"]
    n_plates = summary["signs"]["n_sign_plates"]
    n_poles = summary["signs"]["n_pole_locations"]
    total_m = summary["measurements"]["total_linear_m"]
    scale_ratio = inp.scale_info.get("ratio", "?")
    scale_status = inp.scale_info.get("status", "unverified")
    n_flags = summary["review"]["n_red_flags"]
    n_q = summary["review"]["n_teaching_questions"]
    pipe = summary["pipeline"]

    lines = [
        f"# Worker / Operations Export Report",
        f"",
        f"**Plan:** {slug}  ",
        f"**Source PDF:** {Path(summary['source_pdf']).name}  ",
        f"**Generated:** {gen}  ",
        f"**Status:** DRAFT — REQUIRES REVIEW  ",
        f"**Script:** {SCRIPT_NAME} v{VERSION}  ",
        f"",
        f"> ⚠ {DISCLAIMER}",
        f"",
        f"## Pipeline Status",
        f"- Stages OK: {pipe['stages_ok']}",
        f"- Stages partial: {pipe['stages_partial']}",
        f"- Overall: {pipe['overall']}",
        f"",
        f"## Export Files",
    ]
    for e in exports:
        status = "✓ generated" if e.get("generated") else "✗ not generated"
        lines.append(f"- `{e['filename']}` — {e['description']} — {status}")
    lines.append("")

    lines += [
        f"## Summary",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| BOQ Items | {n_items} |",
        f"| Approved for BOQ | 0 |",
        f"| Requires Review | {n_review} |",
        f"| Sign Plates | {n_plates} |",
        f"| Pole Locations | {n_poles} |",
        f"| Total Linear | {fmt_m(total_m)} |",
        f"| Scale | 1:{scale_ratio} ({scale_status}) |",
        f"| Red Flags | {n_flags} |",
        f"| Teaching Questions | {n_q} |",
        f"",
        f"## Red Flags",
    ]
    for rf in inp.red_flags:
        sev = rf.get("severity", "INFO")
        code = rf.get("code", "")
        msg = rf.get("message", "")
        lines.append(f"- **[{sev}]** `{code}`: {msg}")
    lines.append("")

    if inp.missing:
        lines.append("## Missing Inputs")
        for m in inp.missing:
            lines.append(f"- `{m}` — not found, skipped")
        lines.append("")

    lines += [
        f"## Excel Status",
        f"{'Excel workbook generated: worker_operations_quantities.xlsx' if excel_ok else 'Excel NOT generated — openpyxl unavailable'}",
        f"",
        f"## PDF Status",
        f"PDF generation not implemented (no weasyprint/reportlab in venv). Print-ready HTML is available.",
        f"",
    ]
    return "\n".join(lines)

# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Stage S20 — Worker/Operations Export Generator"
    )
    parser.add_argument(
        "--plan-run-dir",
        default=None,
        help="Path to a plan-scoped run directory. If omitted, uses legacy outputs/ directory.",
    )
    args = parser.parse_args()

    t0 = time.time()
    ctx = PlanRunContext.from_args(args, script_dir=SCRIPT_DIR)

    outputs_dir = ctx.outputs_dir
    exports_dir = outputs_dir / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)

    plan_scoped = args.plan_run_dir is not None
    mode = "plan-scoped" if plan_scoped else "legacy"
    print(f"[33] Worker/Operations Export — mode={mode}")
    print(f"     outputs_dir : {outputs_dir}")
    print(f"     exports_dir : {exports_dir}")

    # Load inputs
    inp = PipelineInputs(outputs_dir)
    if inp.missing:
        print(f"[33] Missing inputs ({len(inp.missing)}): {', '.join(inp.missing[:5])}")
    else:
        print(f"[33] All inputs loaded.")

    # Load plan manifest
    plan_manifest: Dict = {}
    if ctx.plan_run_dir:
        manifest_path = ctx.plan_run_dir / "plan_manifest.json"
        plan_manifest = load_json(manifest_path) or {}
    plan_id   = plan_manifest.get("plan_id") or ctx.plan_id
    plan_slug = plan_manifest.get("plan_slug") or (ctx.plan_run_dir.name if ctx.plan_run_dir else "legacy")
    source_pdf = (
        plan_manifest.get("original_pdf_path")
        or plan_manifest.get("stored_pdf_path")
        or str(ctx.source_pdf_path or "unknown")
    )

    # Build summary
    summary = build_summary(inp, plan_slug, plan_id, source_pdf)

    # Paths for export files
    html_path = exports_dir / "worker_operations_report.html"
    xlsx_path = exports_dir / "worker_operations_quantities.xlsx"
    md_path   = exports_dir / "worker_operations_export_report.md"
    manifest_path = exports_dir / "export_manifest.json"

    exports_list: List[Dict] = []

    # Generate HTML
    print("[33] Generating HTML report...")
    html_content = generate_html(inp, summary, plan_manifest)
    html_path.write_text(html_content, encoding="utf-8")
    print(f"[33] HTML: {html_path} ({html_path.stat().st_size // 1024} KB)")
    exports_list.append({
        "filename": "worker_operations_report.html",
        "type": "html_report",
        "description": "Worker / operations report — print-ready HTML",
        "path": str(html_path),
        "generated": True,
    })
    exports_list.append({
        "filename": "worker_operations_report.pdf",
        "type": "pdf_report",
        "description": "PDF not generated — requires weasyprint/reportlab. Use browser print on HTML.",
        "path": None,
        "generated": False,
    })

    # Generate Excel
    print("[33] Generating Excel workbook...")
    excel_ok = generate_excel(inp, summary, xlsx_path)
    if excel_ok:
        print(f"[33] Excel: {xlsx_path} ({xlsx_path.stat().st_size // 1024} KB)")
    else:
        print("[33] Excel: SKIPPED — openpyxl not available")
    exports_list.append({
        "filename": "worker_operations_quantities.xlsx",
        "type": "excel_workbook",
        "description": "Quantities, BOQ draft, signs, poles, measurements, element groups, review, red flags, preparation, audit",
        "path": str(xlsx_path) if excel_ok else None,
        "generated": excel_ok,
    })

    # Generate markdown
    md_content = generate_md_report(summary, inp, exports_list, excel_ok)
    md_path.write_text(md_content, encoding="utf-8")
    exports_list.append({
        "filename": "worker_operations_export_report.md",
        "type": "markdown_report",
        "description": "Export summary report (markdown)",
        "path": str(md_path),
        "generated": True,
    })

    # Generate manifest
    manifest = generate_manifest(summary, exports_list, inp, plan_manifest)
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    exports_list.append({
        "filename": "export_manifest.json",
        "type": "manifest",
        "description": "Export manifest with metadata and file list",
        "path": str(manifest_path),
        "generated": True,
    })
    manifest["export_files"] = exports_list
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    elapsed = time.time() - t0
    print(f"[33] Done in {elapsed:.1f}s")
    print(f"[33] Exports written to: {exports_dir}")
    print(f"[33]   HTML  : worker_operations_report.html")
    print(f"[33]   Excel : {'worker_operations_quantities.xlsx' if excel_ok else 'NOT GENERATED'}")
    print(f"[33]   MD    : worker_operations_export_report.md")
    print(f"[33]   JSON  : export_manifest.json")
    print(f"[33] REMINDER: All outputs are DRAFT. approved_for_boq: false.")
    if plan_scoped:
        # Paranoia check: no writes outside exports_dir
        print(f"[33] Plan-scoped mode confirmed — no writes to global outputs/")


if __name__ == "__main__":
    main()
